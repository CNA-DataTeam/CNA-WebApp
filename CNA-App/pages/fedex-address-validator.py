"""
address_validation_results.py

Purpose:
    Streamlit page for reviewing FedEx Address Validation results and emailing FedEx.

Behavior:
    - Loads results.csv from configured output path
    - Displays results in a selectable table
    - Allows user to select one or more rows
    - Opens a pre-filled Outlook email addressed to FedEx with row details
    - Preserves global app styling and layout conventions

Inputs:
    - results.csv (path resolved via config)

Outputs:
    - Downloadable Excel file
    - Outlook email draft
"""

# ============================================================
# IMPORTS
# ============================================================
from pathlib import Path
from typing import List, Tuple
from io import BytesIO
from urllib.parse import quote
import webbrowser
from decimal import Decimal, InvalidOperation
import base64

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

import config
import utils

# ============================================================
# PAGE CONFIG (MUST BE FIRST STREAMLIT CALL)
# ============================================================
st.set_page_config(
    page_title="FedEx Address Validator",
    layout="wide",
)

# ============================================================
# GLOBAL STYLING / HEADER
# ============================================================
st.markdown(utils.get_global_css(), unsafe_allow_html=True)

LOGO_PATH = config.LOGO_PATH
logo_b64 = utils.get_logo_base64(str(LOGO_PATH))

st.markdown(
    f"""
    <div class="header-row">
        <img class="header-logo" src="data:image/png;base64,{logo_b64}" />
        <h1 class="header-title">LS - FedEx Address Validator</h1>
    </div>
    """,
    unsafe_allow_html=True,
)

st.divider()


# ============================================================
# CONFIGURATION
# ============================================================
FEDEX_EMAIL_TO = "quickresponse6@fedex.com"
EMAIL_SUBJECT = "Clark National Accounts - Residential Status Dispute"

TABLE_KEY = "results_table"


# ============================================================
# CACHED DATA LOADING
# ============================================================
@st.cache_data
def load_results(file_path: Path) -> pd.DataFrame:
    """Load results file with caching to prevent reloading on every interaction."""
    # CSV reader: tolerate mixed encodings/delimiters and malformed rows.
    last_error: Exception | None = None
    for encoding in ["utf-8-sig", "utf-8", "cp1252", "latin1"]:
        try:
            return pd.read_csv(
                file_path,
                dtype=str,
                keep_default_na=False,
                encoding=encoding,
                sep=None,
                engine="python",
                on_bad_lines="skip",
            )
        except Exception as exc:
            last_error = exc

    raise RuntimeError(f"Unable to read results file: {file_path}") from last_error


def mark_rows_as_disputed(file_path: Path, row_indices: List[int]) -> None:
    """
    Persist Disputed=1 to the source file using original DataFrame indices.
    """
    full_df = load_results(file_path)
    if "Disputed" not in full_df.columns:
        full_df["Disputed"] = ""

    for idx in row_indices:
        if idx in full_df.index:
            full_df.at[idx, "Disputed"] = "1"

    full_df.to_csv(file_path, index=False, encoding="utf-8-sig")

# ============================================================
# LOAD RESULTS
# ============================================================
RESULTS_CSV_FILE: Path = config.ADDRESS_VALIDATION_RESULTS_FILE.with_suffix(".csv")

if not RESULTS_CSV_FILE.exists():
    st.error(f"Results file not found:\n{RESULTS_CSV_FILE}")
    st.stop()

df = load_results(RESULTS_CSV_FILE)

if df.empty:
    st.info("No results available.")
    st.stop()

# ============================================================
# ATTACHMENT DATA BUILDER
# ============================================================
def _first_present(row: pd.Series, candidates: List[str]) -> str:
    for col in candidates:
        if col in row.index:
            value = row.get(col, "")
            if pd.notna(value) and str(value).strip():
                return str(value).strip()
    return ""

def normalize_tracking_number(value: object) -> str:
    """Render scientific-notation tracking numbers as plain strings."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""

    text = str(value).strip()
    if not text:
        return ""

    if "e" in text.lower():
        try:
            number = Decimal(text)
            # FedEx tracking numbers are identifiers; show without exponent.
            return format(number, "f").rstrip("0").rstrip(".")
        except (InvalidOperation, ValueError):
            return text
    return text


def format_currency_display(value: object) -> str:
    """Render numeric values as currency for table display."""
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return ""
    return f"${num:,.2f}"


def build_attachment_df(rows: pd.DataFrame) -> pd.DataFrame:
    dispute_reason = "Address is classified as business. Residential fee would not apply."

    attachment_df = pd.DataFrame(
        {
            "Account": rows.apply(
                lambda r: _first_present(
                    r,
                    [
                        "Bill to Account Number",
                        "Account Number",
                        "AccountNumber",
                        "Account",
                        "Billed Account",
                    ],
                ),
                axis=1,
            ),
            "Invoice": rows.apply(
                lambda r: _first_present(r, ["Invoice", "InvoiceNumber", "Invoice Number"]),
                axis=1,
            ),
            "Tracking Number": rows.apply(
                lambda r: normalize_tracking_number(
                    _first_present(r, ["Tracking Number", "InvTrackingNumber", "Tracking"])
                ),
                axis=1,
            ),
            "Amount Billed": pd.to_numeric(rows.get("Net Charge Amount"), errors="coerce"),
            "Credit Requested": 2.38,
            "Reason": dispute_reason,
        }
    )
    return attachment_df


def create_excel_download(rows: pd.DataFrame) -> Tuple[str, bytes]:
    attachment_df = build_attachment_df(rows)
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"fedex_residential_dispute_{timestamp}.xlsx"

    base_buffer = BytesIO()
    attachment_df.to_excel(base_buffer, index=False)
    base_buffer.seek(0)

    # Apply currency formatting to Amount Billed and Credit Requested columns.
    wb = load_workbook(base_buffer)
    ws = wb.active
    currency_cols = [4, 5]  # 1-based indexes in final sheet
    for col_idx in currency_cols:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "$#,##0.00"

    formatted_buffer = BytesIO()
    wb.save(formatted_buffer)
    formatted_buffer.seek(0)

    return file_name, formatted_buffer.getvalue()

def trigger_file_download(file_name: str, file_bytes: bytes) -> None:
    """Trigger a browser download in the same click cycle."""
    safe_name = file_name.replace('"', "_")
    file_b64 = base64.b64encode(file_bytes).decode("ascii")
    components.html(
        f"""
        <a id="dl" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{file_b64}" download="{safe_name}"></a>
        <script>
            const a = document.getElementById("dl");
            if (a) a.click();
        </script>
        """,
        height=0,
    )


# ============================================================
# EMAIL DISPATCH (OUTLOOK ONLY)
# ============================================================
def open_email(to_addr: str, subject: str, body: str) -> Tuple[bool, str]:
    # Prefer default mail handler (New Outlook when configured as default).
    mailto = f"mailto:{quote(to_addr)}?subject={quote(subject)}&body={quote(body)}"
    try:
        if webbrowser.open(mailto):
            return True, "Opened via default mail app."
    except Exception:
        pass

    com_initialized = False
    try:
        import pythoncom  # type: ignore
        import win32com.client as win32  # type: ignore

        pythoncom.CoInitialize()
        com_initialized = True

        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = to_addr
        mail.Subject = subject
        mail.Body = body
        mail.Display()
        return True, "Opened via Outlook desktop (classic)."

    except ModuleNotFoundError:
        return False, "Outlook automation library (pywin32) is not installed."
    except Exception as exc:
        return False, f"Outlook automation failed ({type(exc).__name__}): {exc}"
    finally:
        if com_initialized:
            try:
                pythoncom.CoUninitialize()  # type: ignore[name-defined]
            except Exception:
                pass

# ============================================================
# COLUMN FILTERING AND RENAMING
# ============================================================
# Convert InvoiceDate format if it exists (from yyyymmdd to mm/dd/yyyy)
if "InvoiceDate" in df.columns:
    df["InvoiceDate"] = pd.to_datetime(df["InvoiceDate"], format="%Y%m%d", errors="coerce").dt.strftime("%m/%d/%Y")

# Show only non-disputed rows (empty Disputed values).
if "Disputed" in df.columns:
    disputed_values = df["Disputed"].fillna("").astype(str).str.strip()
    df = df[disputed_values.eq("")]

# Build a display-only dataframe. Keep `df` intact for downstream actions like Excel generation.
display_df = df.copy()

# Remove unwanted display columns.
columns_to_remove = [
    "StreetLine1",
    "StreetLine2",
    "PostalCode",
    "Shipment Date",
    "OriginalCustomerReference",
    "Original Customer Reference",
    "ID",
    "CountryCode",
    "Transportation Charge Amount",
    "Disputed",
    "City",
]
columns_to_remove.extend([col for col in display_df.columns if "Recipient" in col or "recipient" in col])
display_df = display_df.drop(columns=[col for col in columns_to_remove if col in display_df.columns])

# Rename display columns.
display_df = display_df.rename(
    columns={
        "InvTrackingNumber": "Tracking Number",
        "ResidentialStatusMatch": "Residential Match",
    }
)
if "StateOrProvince" in display_df.columns:
    if "State" in display_df.columns:
        display_df = display_df.drop(columns=["StateOrProvince"])
    else:
        display_df = display_df.rename(columns={"StateOrProvince": "State"})

# Ensure tracking values display in full (not scientific notation).
tracking_col = next((c for c in ["Tracking Number", "InvTrackingNumber", "Tracking"] if c in display_df.columns), None)
if tracking_col:
    display_df[tracking_col] = display_df[tracking_col].apply(normalize_tracking_number)

# Format Net Charge Amount as currency for display.
if "Net Charge Amount" in display_df.columns:
    display_df["Net Charge Amount"] = display_df["Net Charge Amount"].apply(format_currency_display)

# Move Classification before Match Type / Residential Match.
classification_col = next((c for c in ["Classification", "classification"] if c in display_df.columns), None)
match_type_col = next(
    (c for c in ["Match Type", "MatchType", "Residential Match"] if c in display_df.columns),
    None,
)
if classification_col and match_type_col and classification_col != match_type_col:
    cols = list(display_df.columns)
    cols.remove(classification_col)
    match_idx = cols.index(match_type_col)
    cols.insert(match_idx, classification_col)
    display_df = display_df[cols]

# ============================================================
# FILTERS
# ============================================================
with st.expander("Filters", expanded=True):
    state_col = next(
        (c for c in ["State", "StateOrProvince", "Recipient State"] if c in display_df.columns),
        None,
    )
    service_col = next(
        (c for c in ["Service Type", "ServiceType", "Service"] if c in display_df.columns),
        None,
    )
    invoice_date_col = next(
        (c for c in ["InvoiceDate", "Invoice Date", "Invoice_Date"] if c in display_df.columns),
        None,
    )

    left_col, right_col = st.columns(2)
    with left_col:
        status_filter = st.multiselect(
            "Residential Match",
            sorted(display_df["Residential Match"].dropna().astype(str).unique().tolist()),
        )
        state_options = sorted(display_df[state_col].dropna().astype(str).unique().tolist()) if state_col else []
        state_filter = st.multiselect("State", state_options, disabled=state_col is None)

    with right_col:
        if invoice_date_col:
            invoice_dates = pd.to_datetime(df[invoice_date_col], errors="coerce").dropna()
            if invoice_dates.empty:
                invoice_date_range = st.date_input(
                    "Invoice Date",
                    disabled=True,
                    help="No valid invoice dates found in data.",
                )
            else:
                min_date = invoice_dates.min().date()
                max_date = invoice_dates.max().date()
                invoice_date_range = st.date_input(
                    "Invoice Date",
                    value=(min_date, max_date),
                    min_value=min_date,
                    max_value=max_date,
                )
        else:
            invoice_date_range = st.date_input(
                "Invoice Date",
                disabled=True,
                help="Invoice Date column not found.",
            )
        service_options = sorted(display_df[service_col].dropna().astype(str).unique().tolist()) if service_col else []
        service_filter = st.multiselect("Service Type", service_options, disabled=service_col is None)

view_df = display_df
if state_filter and state_col:
    view_df = view_df[view_df[state_col].fillna("").astype(str).isin(state_filter)]
if service_filter and service_col:
    view_df = view_df[view_df[service_col].fillna("").astype(str).isin(service_filter)]
if invoice_date_col:
    if isinstance(invoice_date_range, (tuple, list)) and len(invoice_date_range) == 2:
        start_date, end_date = invoice_date_range
    else:
        start_date = invoice_date_range
        end_date = invoice_date_range

    if start_date and end_date:
        invoice_dates = pd.to_datetime(view_df[invoice_date_col], errors="coerce").dt.date
        view_df = view_df[(invoice_dates >= start_date) & (invoice_dates <= end_date)]
if status_filter:
    view_df = view_df[view_df["Residential Match"].isin(status_filter)]


# ============================================================
# INITIALIZE SESSION STATE FOR SELECTIONS
# ============================================================
if "select_all" not in st.session_state:
    st.session_state.select_all = False

if "editor_df" not in st.session_state:
    st.session_state.editor_df = None
if "apply_select_all" not in st.session_state:
    st.session_state.apply_select_all = False

# ============================================================
# TOGGLE SELECT / SEND EMAIL BUTTONS
# ============================================================
col1, col2, col3, col4, col5 = st.columns([1.5, 5.1, 2.4, 2.6, 2.4], width="stretch")

with col1:
    # Determine button label based on current state
    if st.session_state.select_all:
        button_label = "Deselect All ❌"
    else:
        button_label = "Select All ✅"
    
    if st.button(button_label, key="toggle_select"):
        st.session_state.select_all = not st.session_state.select_all
        st.session_state.apply_select_all = True
        st.rerun()

# col2 is empty spacer in the middle

with col3:
    # This will be populated after we know selected_rows
    generate_button_placeholder = st.empty()

with col4:
    # This will be populated after we know selected_rows
    email_button_placeholder = st.empty()

with col5:
    # This will be populated after we know selected_rows
    disputed_button_placeholder = st.empty()

# ============================================================
# TABLE WITH CHECKBOX SELECTION
# ============================================================

# Build a stable display DF with a persisted Select column.
base_df = view_df.copy()

# Seed editor_df the first time, or when shape/columns change (filters, new data, etc.)
needs_reset = (
    st.session_state.editor_df is None
    or len(st.session_state.editor_df) != len(base_df)
    or list(st.session_state.editor_df.columns) != (["Select"] + list(base_df.columns))
)

if needs_reset:
    seed_df = base_df.copy()
    seed_df.insert(0, "Select", False)
    st.session_state.editor_df = seed_df

# Apply Select All / Deselect All exactly once when user clicks the toggle button.
if st.session_state.apply_select_all:
    st.session_state.editor_df["Select"] = bool(st.session_state.select_all)
    st.session_state.apply_select_all = False

edited_df = st.data_editor(
    st.session_state.editor_df,
    use_container_width=True,
    hide_index=True,
    height=700,
    key="address_validation_editor",
    column_config={
        "Select": st.column_config.CheckboxColumn(
            "Select",
            help="Select rows to include in the email",
        )
    },
    disabled=[c for c in st.session_state.editor_df.columns if c != "Select"],
)

# Persist edits so selections survive reruns (button clicks cause reruns).
st.session_state.editor_df = edited_df

# If user manually changes any checkbox while select_all is True, drop select_all flag
# (this prevents the UI label from lying).
if st.session_state.select_all and edited_df["Select"].sum() != len(edited_df):
    st.session_state.select_all = False

# ============================================================
# SELECTION EXTRACTION
# ============================================================
selected_mask = edited_df["Select"]
selected_indices = edited_df.index[selected_mask]
has_selection = len(selected_indices) > 0
selected_rows_for_email = df.loc[selected_indices]
# ============================================================
# ACTION BUTTONS (in right columns)
# ============================================================
with generate_button_placeholder:
    generate_dispute_clicked = st.button(
        "Generate Dispute File",
        disabled=not has_selection,
        key="generate_dispute",
    )

with email_button_placeholder:
    send_email_clicked = st.button(
        "Send Email to FedEx 📨",
        disabled=not has_selection,
        key="send_email",
    )

with disputed_button_placeholder:
    mark_disputed_clicked = st.button(
        "Mark as Disputed",
        disabled=not has_selection,
        key="mark_disputed",
    )

if generate_dispute_clicked:
    try:
        with st.spinner("Creating dispute Excel file..."):
            file_name, file_bytes = create_excel_download(selected_rows_for_email)
        trigger_file_download(file_name, file_bytes)
        st.success("Dispute file generated and download started.")
        st.download_button(
            "If download did not start, click to download",
            data=file_bytes,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"fallback_download_{file_name}",
        )
    except Exception as exc:
        st.error(f"Generate dispute file failed: {type(exc).__name__}: {exc}")

if send_email_clicked:
    try:
        if not FEDEX_EMAIL_TO.strip():
            st.error("FedEx recipient email is blank. Set FEDEX_EMAIL_TO first.")
        else:
            email_body = (
                "Hello FedEx Team,\n\n"
                "Please review the selected shipments in the dispute file.\n\n"
                "Thank you,"
            )
            email_opened, email_status = open_email(FEDEX_EMAIL_TO, EMAIL_SUBJECT, email_body)
            if email_opened:
                st.success("Outlook draft opened.")
            else:
                st.error("Failed to open editable Outlook draft.")
                st.caption(email_status)
    except Exception as exc:
        st.error(f"Send email failed: {type(exc).__name__}: {exc}")

if mark_disputed_clicked:
    try:
        mark_rows_as_disputed(RESULTS_CSV_FILE, selected_indices.tolist())
        load_results.clear()
        st.success("Selected rows marked as disputed.")
        st.rerun()
    except Exception as exc:
        st.error(f"Failed to mark rows as disputed: {exc}")

# ============================================================
# FOOTNOTE
# ============================================================
st.caption(
    "Emails open as editable drafts in Outlook desktop. Download the dispute Excel separately."
)
