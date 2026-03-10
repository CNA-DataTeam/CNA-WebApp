"""
address_validation_results.py

Purpose:
    Streamlit page for reviewing FedEx Address Validation results.

Behavior:
    - Loads results.csv from configured output path
    - Displays results in a selectable table
    - Supports filtering for analysis
    - Supports dispute workflow actions (export/email/mark disputed)
    - Preserves global app styling and layout conventions

Inputs:
    - results.csv (path resolved via config)
"""

# ============================================================
# IMPORTS
# ============================================================
from pathlib import Path
from typing import List, Tuple
from io import BytesIO
from urllib.parse import quote
import webbrowser
from decimal import Decimal, InvalidOperation
import base64

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

import config
import utils

LOGGER = utils.get_page_logger("FedEx Address Validator")
PAGE_TITLE = utils.get_registry_page_title(__file__, "FedEx Address Validator")

# ============================================================
# PAGE CONFIG (MUST BE FIRST STREAMLIT CALL)
# ============================================================
st.set_page_config(
    page_title=PAGE_TITLE,
    layout="wide",
)
utils.render_app_logo()
utils.log_page_open_once("fedex_address_validator_page", LOGGER)
if "_fedex_render_logged" not in st.session_state:
    st.session_state._fedex_render_logged = True
    LOGGER.info("Render UI.")

# ============================================================
# GLOBAL STYLING / HEADER
# ============================================================
st.markdown(utils.get_global_css(), unsafe_allow_html=True)

utils.render_page_header(PAGE_TITLE)


# ============================================================
# CONFIGURATION
# ============================================================
FEDEX_EMAIL_TO = "quickresponse6@fedex.com"
EMAIL_SUBJECT = "Clark National Accounts - Residential Status Dispute"
ROW_ID_COL = "__source_row_id"
BASE_RESIDENTIAL_MATCH_VALUES = {"mismatch", "mixed"}
BASE_RESIDENTIAL_MATCH_COLUMNS = ("ResidentialStatusMatch", "Residential Match")
DISPUTE_AMOUNT_COLUMNS = ("Dispute Amount", "DisputeAmount")
TABLE_INTRO_CAPTION = (
    'The following Tracking Numbers were charged a "Residential Delivery fee"; they were checked '
    "using Web's Address Validation API, and the API did not confirm a residential status for the "
    "address they shipped to."
)

if "fedex_mark_disputed_confirm_open" not in st.session_state:
    st.session_state.fedex_mark_disputed_confirm_open = False
if "fedex_mark_disputed_confirm_rendered" not in st.session_state:
    st.session_state.fedex_mark_disputed_confirm_rendered = False
if "fedex_mark_disputed_row_ids" not in st.session_state:
    st.session_state.fedex_mark_disputed_row_ids = []
if "fedex_mark_disputed_success" not in st.session_state:
    st.session_state.fedex_mark_disputed_success = False


# ============================================================
# CACHED DATA LOADING
# ============================================================
@st.cache_data
def load_results(file_path: Path) -> pd.DataFrame:
    """Load results file with caching to prevent reloading on every interaction."""
    # CSV reader: tolerate mixed encodings/delimiters and malformed rows.
    last_error: Exception | None = None
    for encoding in ["utf-8-sig", "utf-8", "cp1252", "latin1"]:
        try:
            return pd.read_csv(
                file_path,
                dtype=str,
                keep_default_na=False,
                encoding=encoding,
                sep=None,
                engine="python",
                on_bad_lines="skip",
            )
        except Exception as exc:
            last_error = exc

    raise RuntimeError(f"Unable to read results file: {file_path}") from last_error


def mark_rows_as_disputed(file_path: Path, row_indices: List[object]) -> None:
    """
    Persist Disputed=1 to the source file using original DataFrame indices.
    """
    full_df = load_results(file_path)
    if "Disputed" not in full_df.columns:
        full_df["Disputed"] = ""

    normalized_indices: List[int] = []
    for raw_idx in row_indices:
        try:
            normalized_indices.append(int(raw_idx))
        except (TypeError, ValueError):
            LOGGER.warning("Skipping invalid row index while marking disputed: %r", raw_idx)

    for idx in normalized_indices:
        if idx in full_df.index:
            full_df.at[idx, "Disputed"] = "1"

    full_df.to_csv(file_path, index=False, encoding="utf-8-sig")

# ============================================================
# LOAD RESULTS
# ============================================================
RESULTS_CSV_FILE: Path = config.ADDRESS_VALIDATION_RESULTS_FILE.with_suffix(".csv")

if not RESULTS_CSV_FILE.exists():
    LOGGER.error("Results file not found: %s", RESULTS_CSV_FILE)
    st.error(f"Results file not found:\n{RESULTS_CSV_FILE}")
    st.stop()

df = load_results(RESULTS_CSV_FILE)
if "_fedex_rows_logged" not in st.session_state:
    st.session_state._fedex_rows_logged = True
    LOGGER.info("Loaded FedEx validation results | rows=%s file='%s'", len(df), RESULTS_CSV_FILE)

if df.empty:
    LOGGER.info("Results file is present but contains no rows.")
    st.info("No results available.")
    st.stop()

# ============================================================
# DISPLAY HELPERS
# ============================================================
def normalize_tracking_number(value: object) -> str:
    """Render scientific-notation tracking numbers as plain strings."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""

    text = str(value).strip()
    if not text:
        return ""

    if "e" in text.lower():
        try:
            number = Decimal(text)
            # FedEx tracking numbers are identifiers; show without exponent.
            return format(number, "f").rstrip("0").rstrip(".")
        except (InvalidOperation, ValueError):
            return text
    return text


def format_currency_display(value: object) -> str:
    """Render numeric values as currency for table display."""
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return ""
    return f"${num:,.2f}"


def sum_currency_values(rows: pd.DataFrame, candidates: Tuple[str, ...]) -> float:
    """Sum the first matching currency-like column from the provided dataframe."""
    currency_col = next((col for col in candidates if col in rows.columns), None)
    if not currency_col:
        return 0.0
    return float(pd.to_numeric(rows[currency_col], errors="coerce").fillna(0).sum())


def _first_present(row: pd.Series, candidates: List[str]) -> str:
    for col in candidates:
        if col in row.index:
            value = row.get(col, "")
            if pd.notna(value) and str(value).strip():
                return str(value).strip()
    return ""


def build_attachment_df(rows: pd.DataFrame) -> pd.DataFrame:
    dispute_reason = "Address is classified as business. Residential fee would not apply."

    attachment_df = pd.DataFrame(
        {
            "Account": rows.apply(
                lambda r: _first_present(
                    r,
                    [
                        "Bill to Account Number",
                        "Account Number",
                        "AccountNumber",
                        "Account",
                        "Billed Account",
                    ],
                ),
                axis=1,
            ),
            "Invoice": rows.apply(
                lambda r: _first_present(r, ["Invoice", "InvoiceNumber", "Invoice Number"]),
                axis=1,
            ),
            "Tracking Number": rows.apply(
                lambda r: normalize_tracking_number(
                    _first_present(r, ["Tracking Number", "InvTrackingNumber", "Tracking"])
                ),
                axis=1,
            ),
            "Amount Billed": pd.to_numeric(rows.get("Net Charge Amount"), errors="coerce"),
            "Credit Requested": 2.38,
            "Reason": dispute_reason,
        }
    )
    return attachment_df


def create_excel_download(rows: pd.DataFrame) -> Tuple[str, bytes]:
    attachment_df = build_attachment_df(rows)
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"fedex_residential_dispute_{timestamp}.xlsx"

    base_buffer = BytesIO()
    attachment_df.to_excel(base_buffer, index=False)
    base_buffer.seek(0)

    # Apply currency formatting to Amount Billed and Credit Requested columns.
    wb = load_workbook(base_buffer)
    ws = wb.active
    currency_cols = [4, 5]  # 1-based indexes in final sheet
    for col_idx in currency_cols:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "$#,##0.00"

    formatted_buffer = BytesIO()
    wb.save(formatted_buffer)
    formatted_buffer.seek(0)

    return file_name, formatted_buffer.getvalue()


def trigger_file_download(file_name: str, file_bytes: bytes) -> None:
    """Trigger a browser download in the same click cycle."""
    safe_name = file_name.replace('"', "_")
    file_b64 = base64.b64encode(file_bytes).decode("ascii")
    components.html(
        f"""
        <a id="dl" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{file_b64}" download="{safe_name}"></a>
        <script>
            const a = document.getElementById("dl");
            if (a) a.click();
        </script>
        """,
        height=0,
    )


def open_email(to_addr: str, subject: str, body: str) -> Tuple[bool, str]:
    # Prefer default mail handler (New Outlook when configured as default).
    mailto = f"mailto:{quote(to_addr)}?subject={quote(subject)}&body={quote(body)}"
    try:
        if webbrowser.open(mailto):
            return True, "Opened via default mail app."
    except Exception:
        pass

    com_initialized = False
    try:
        import pythoncom  # type: ignore
        import win32com.client as win32  # type: ignore

        pythoncom.CoInitialize()
        com_initialized = True

        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = to_addr
        mail.Subject = subject
        mail.Body = body
        mail.Display()
        return True, "Opened via Outlook desktop (classic)."

    except ModuleNotFoundError:
        return False, "Outlook automation library (pywin32) is not installed."
    except Exception as exc:
        return False, f"Outlook automation failed ({type(exc).__name__}): {exc}"
    finally:
        if com_initialized:
            try:
                pythoncom.CoUninitialize()  # type: ignore[name-defined]
            except Exception:
                pass


@st.dialog("Confirm")
def confirm_mark_as_disputed_dialog() -> None:
    """Prompt before persisting the disputed flag for the visible rows."""
    st.write("Are you sure you want to mark these orders as disputed?")

    confirm_col, cancel_col = st.columns(2)
    with confirm_col:
        if st.button(
            "Confirm",
            type="primary",
            width="stretch",
            key="fedex_confirm_mark_disputed_button",
        ):
            pending_row_ids = st.session_state.get("fedex_mark_disputed_row_ids", [])
            try:
                mark_rows_as_disputed(RESULTS_CSV_FILE, pending_row_ids)
                load_results.clear()
                LOGGER.info("Marked %s displayed row(s) as disputed.", len(pending_row_ids))
                st.session_state.fedex_mark_disputed_success = True
                st.session_state.fedex_mark_disputed_confirm_open = False
                st.session_state.fedex_mark_disputed_confirm_rendered = False
                st.session_state.fedex_mark_disputed_row_ids = []
                st.rerun()
            except Exception as exc:
                LOGGER.exception("Failed to mark rows as disputed: %s", exc)
                st.error(f"Failed to mark rows as disputed: {exc}")
    with cancel_col:
        if st.button("Cancel", width="stretch", key="fedex_cancel_mark_disputed_button"):
            st.session_state.fedex_mark_disputed_confirm_open = False
            st.session_state.fedex_mark_disputed_confirm_rendered = False
            st.session_state.fedex_mark_disputed_row_ids = []
            st.rerun()

# ============================================================
# COLUMN FILTERING AND RENAMING
# ============================================================
# Convert InvoiceDate format if it exists (from yyyymmdd to mm/dd/yyyy)
if "InvoiceDate" in df.columns:
    df["InvoiceDate"] = pd.to_datetime(df["InvoiceDate"], format="%Y%m%d", errors="coerce").dt.strftime("%m/%d/%Y")

# Show only non-disputed rows (empty Disputed values).
if "Disputed" in df.columns:
    disputed_values = df["Disputed"].fillna("").astype(str).str.strip()
    df = df[disputed_values.eq("")]

# Keep only rows with Residential Match statuses that need review.
residential_match_col = next((c for c in BASE_RESIDENTIAL_MATCH_COLUMNS if c in df.columns), None)
if residential_match_col:
    base_status_values = df[residential_match_col].fillna("").astype(str).str.strip().str.lower()
    df = df[base_status_values.isin(BASE_RESIDENTIAL_MATCH_VALUES)]
else:
    LOGGER.warning(
        "Residential Match column not found. Expected one of: %s",
        ", ".join(BASE_RESIDENTIAL_MATCH_COLUMNS),
    )

# Persist source row identity for stable selection mapping.
df = df.copy()
df[ROW_ID_COL] = df.index.astype(str)

# Build a display-only dataframe for the table.
display_df = df.copy()

# Remove unwanted display columns.
columns_to_remove = [
    "StreetLine1",
    "StreetLine2",
    "PostalCode",
    "State",
    "StateOrProvince",
    "Shipment Date",
    "OriginalCustomerReference",
    "Original Customer Reference",
    "ID",
    "CountryCode",
    "Transportation Charge Amount",
    "Disputed",
    "City",
]
columns_to_remove.extend([col for col in display_df.columns if "Recipient" in col or "recipient" in col])
display_df = display_df.drop(columns=[col for col in columns_to_remove if col in display_df.columns])

# Rename display columns.
display_df = display_df.rename(
    columns={
        "InvTrackingNumber": "Tracking Number",
        "ResidentialStatusMatch": "Residential Match",
    }
)

# Ensure tracking values display in full (not scientific notation).
tracking_col = next((c for c in ["Tracking Number", "InvTrackingNumber", "Tracking"] if c in display_df.columns), None)
if tracking_col:
    display_df[tracking_col] = display_df[tracking_col].apply(normalize_tracking_number)

# Format Net Charge Amount as currency for display.
if "Net Charge Amount" in display_df.columns:
    display_df["Net Charge Amount"] = display_df["Net Charge Amount"].apply(format_currency_display)
if "Dispute Amount" in display_df.columns:
    display_df["Dispute Amount"] = display_df["Dispute Amount"].apply(format_currency_display)

# Move Classification before Match Type / Residential Match.
classification_col = next((c for c in ["Classification", "classification"] if c in display_df.columns), None)
match_type_col = next(
    (c for c in ["Match Type", "MatchType", "Residential Match"] if c in display_df.columns),
    None,
)
if classification_col and match_type_col and classification_col != match_type_col:
    cols = list(display_df.columns)
    cols.remove(classification_col)
    match_idx = cols.index(match_type_col)
    cols.insert(match_idx, classification_col)
    display_df = display_df[cols]

# ============================================================
# FILTERS
# ============================================================
with st.expander("Filters", expanded=True):
    service_col = next(
        (c for c in ["Service Type", "ServiceType", "Service"] if c in display_df.columns),
        None,
    )
    invoice_date_col = next(
        (c for c in ["InvoiceDate", "Invoice Date", "Invoice_Date"] if c in display_df.columns),
        None,
    )
    residential_match_options = (
        sorted(display_df["Residential Match"].dropna().astype(str).unique().tolist())
        if "Residential Match" in display_df.columns
        else []
    )
    classification_options = (
        sorted(display_df[classification_col].dropna().astype(str).unique().tolist())
        if classification_col
        else []
    )

    top_left_col, top_right_col = st.columns(2)
    with top_left_col:
        status_filter = st.multiselect(
            "Residential Match",
            residential_match_options,
            disabled="Residential Match" not in display_df.columns,
            key="fedex_residential_match_filter",
        )
    with top_right_col:
        classification_filter = st.multiselect(
            "Classification",
            classification_options,
            disabled=classification_col is None,
            key="fedex_classification_filter",
        )

    bottom_left_col, bottom_right_col = st.columns(2)
    with bottom_left_col:
        service_options = sorted(display_df[service_col].dropna().astype(str).unique().tolist()) if service_col else []
        service_filter = st.multiselect("Service Type", service_options, disabled=service_col is None)

    with bottom_right_col:
        if invoice_date_col:
            invoice_dates = pd.to_datetime(df[invoice_date_col], errors="coerce").dropna()
            if invoice_dates.empty:
                invoice_date_range = st.date_input(
                    "Invoice Date",
                    disabled=True,
                    help="No valid invoice dates found in data.",
                )
            else:
                min_date = invoice_dates.min().date()
                max_date = invoice_dates.max().date()
                invoice_date_range = st.date_input(
                    "Invoice Date",
                    value=(min_date, max_date),
                    min_value=min_date,
                    max_value=max_date,
                )
        else:
            invoice_date_range = st.date_input(
                "Invoice Date",
                disabled=True,
                help="Invoice Date column not found.",
            )

view_df = display_df
if service_filter and service_col:
    view_df = view_df[view_df[service_col].fillna("").astype(str).isin(service_filter)]
if classification_filter and classification_col:
    view_df = view_df[view_df[classification_col].fillna("").astype(str).isin(classification_filter)]
if invoice_date_col:
    if isinstance(invoice_date_range, (tuple, list)) and len(invoice_date_range) == 2:
        start_date, end_date = invoice_date_range
    else:
        start_date = invoice_date_range
        end_date = invoice_date_range

    if start_date and end_date:
        invoice_dates = pd.to_datetime(view_df[invoice_date_col], errors="coerce").dt.date
        view_df = view_df[(invoice_dates >= start_date) & (invoice_dates <= end_date)]
if status_filter:
    view_df = view_df[view_df["Residential Match"].isin(status_filter)]

_start_date_text = str(start_date) if "start_date" in locals() else ""
_end_date_text = str(end_date) if "end_date" in locals() else ""
filter_signature = (
    tuple(sorted(status_filter)),
    tuple(sorted(classification_filter)),
    tuple(sorted(service_filter)),
    _start_date_text,
    _end_date_text,
)
if st.session_state.get("_fedex_filter_signature") != filter_signature:
    st.session_state._fedex_filter_signature = filter_signature
    LOGGER.info(
        "Filters updated | status=%s classification=%s services=%s date_range=%s..%s",
        len(status_filter),
        len(classification_filter),
        len(service_filter),
        _start_date_text,
        _end_date_text,
    )
if st.session_state.get("_fedex_last_view_rows") != len(view_df):
    st.session_state._fedex_last_view_rows = len(view_df)
    LOGGER.info("Rows visible after filter | rows=%s", len(view_df))


# ============================================================
# TABLE CONTROLS / DISPLAY
# ============================================================
visible_row_ids = view_df[ROW_ID_COL].astype(str).tolist() if ROW_ID_COL in view_df.columns else []
visible_rows_for_actions = (
    df.set_index(ROW_ID_COL, drop=False).loc[visible_row_ids] if visible_row_ids else df.head(0)
)
visible_source_indices: List[int] = []
for row_id in visible_row_ids:
    try:
        visible_source_indices.append(int(row_id))
    except (TypeError, ValueError):
        LOGGER.warning("Skipping invalid visible row id: %r", row_id)

table_df = view_df.drop(columns=[ROW_ID_COL], errors="ignore")
total_dispute_amount = sum_currency_values(visible_rows_for_actions, DISPUTE_AMOUNT_COLUMNS)

if st.session_state.get("fedex_mark_disputed_success"):
    st.success("Displayed rows marked as disputed.")
    st.session_state.fedex_mark_disputed_success = False

st.caption(TABLE_INTRO_CAPTION)

summary_col1, summary_col2, summary_spacer, action_col = st.columns([1.8, 2.2, 3.2, 4.8], width="stretch")
with summary_col1:
    st.metric("Orders", f"{len(table_df):,}")
with summary_col2:
    st.metric("Total Amount to Dispute", format_currency_display(total_dispute_amount))
with action_col:
    action_btn1, action_btn2, action_btn3 = st.columns(3)
    with action_btn1:
        generate_dispute_clicked = st.button("Generate Dispute File", key="generate_dispute", width="stretch")
    with action_btn2:
        send_email_clicked = st.button("Send Email to FedEx", key="send_email", width="stretch")
    with action_btn3:
        mark_disputed_clicked = st.button("Mark as Disputed", key="mark_disputed", width="stretch")

st.dataframe(
    table_df,
    use_container_width=True,
    hide_index=True,
    height=700,
)

if generate_dispute_clicked:
    LOGGER.info("Decision: Generate Dispute File clicked.")
    try:
        if visible_rows_for_actions.empty:
            st.warning("No rows are currently displayed to export.")
        else:
            with st.spinner("Creating dispute Excel file..."):
                file_name, file_bytes = create_excel_download(visible_rows_for_actions)
            trigger_file_download(file_name, file_bytes)
            LOGGER.info("Generated dispute file '%s' for %s displayed row(s).", file_name, len(visible_rows_for_actions))
            st.success("Dispute file generated and download started.")
            st.download_button(
                "If download did not start, click to download",
                data=file_bytes,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"fallback_download_{file_name}",
            )
    except Exception as exc:
        LOGGER.exception("Generate dispute file failed: %s", exc)
        st.error(f"Generate dispute file failed: {type(exc).__name__}: {exc}")

if send_email_clicked:
    LOGGER.info("Decision: Send Email to FedEx clicked.")
    try:
        if visible_rows_for_actions.empty:
            st.warning("No rows are currently displayed to email.")
        elif not FEDEX_EMAIL_TO.strip():
            LOGGER.error("FedEx recipient email is blank.")
            st.error("FedEx recipient email is blank. Set FEDEX_EMAIL_TO first.")
        else:
            email_body = (
                "Hello FedEx Team,\n\n"
                "Please review the displayed shipments in the dispute file.\n\n"
                "Thank you,"
            )
            email_opened, email_status = open_email(FEDEX_EMAIL_TO, EMAIL_SUBJECT, email_body)
            if email_opened:
                LOGGER.info("Opened Outlook draft for %s displayed row(s).", len(visible_rows_for_actions))
                st.success("Outlook draft opened.")
            else:
                LOGGER.warning("Failed to open editable Outlook draft. status=%s", email_status)
                st.error("Failed to open editable Outlook draft.")
                st.caption(email_status)
    except Exception as exc:
        LOGGER.exception("Send email failed: %s", exc)
        st.error(f"Send email failed: {type(exc).__name__}: {exc}")

if mark_disputed_clicked:
    LOGGER.info("Decision: Mark as Disputed clicked.")
    if not visible_source_indices:
        st.warning("No rows are currently displayed to mark as disputed.")
    else:
        st.session_state.fedex_mark_disputed_row_ids = visible_source_indices
        st.session_state.fedex_mark_disputed_confirm_open = True
        st.session_state.fedex_mark_disputed_confirm_rendered = False

if (
    st.session_state.fedex_mark_disputed_confirm_open
    and not st.session_state.fedex_mark_disputed_confirm_rendered
):
    st.session_state.fedex_mark_disputed_confirm_rendered = True
    confirm_mark_as_disputed_dialog()

st.caption("Emails open as editable drafts in Outlook desktop. Download the dispute Excel separately.")

